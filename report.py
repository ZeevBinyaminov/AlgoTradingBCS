from datetime import datetime
from os import path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


class Report:
    report_file = 'report.xlsx'
    headings = ['SEC_CODE', 'CLASS_CODE', 'QUANTITY', 'PRICE', 'AMOUNT', 'VWAP']
    headings_letter = ['A', 'B', 'C', 'D', 'E', 'F']
    TODAY = datetime.today().strftime("%d.%m.%Y")

    def __init__(self, filename=report_file):
        self.filename: str = filename
        if not path.exists(self.filename):
            self.wb = Workbook()
            self.wb.save(self.filename)
        self.wb: Workbook = load_workbook(filename)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.filename:
            self.save()
        # return True

    def save(self):
        self.wb.save(self.filename)

    def create_current_report(self) -> bool:
        if self.TODAY not in self.wb.sheetnames:
            self.wb.create_sheet(self.TODAY)
            ws = self.wb[self.TODAY]
            for i in range(1, len(self.headings) + 1):
                ws.cell(1, i, self.headings[i - 1])
                ws.cell(1, i).alignment = Alignment(horizontal='center')

            for letter in self.headings_letter:
                ws.column_dimensions[letter].width = 20

            self.wb.save(self.filename)
            return True
        return False

    @property
    def current_report_sheet(self):
        self.create_current_report()
        return self.wb[self.TODAY]
